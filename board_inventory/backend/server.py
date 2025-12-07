from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, validator
from typing import List, Optional
import uuid
from datetime import datetime, timedelta, timezone
import jwt
import bcrypt
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Permission system constants
AVAILABLE_PERMISSIONS = [
    "view_dashboard",
    "view_categories", 
    "create_categories",
    "edit_categories",
    "view_boards",
    "create_boards", 
    "edit_boards",
    "view_inward",
    "create_inward",
    "view_search",
    "view_issue_requests",
    "create_issue_requests",
    "approve_issue_requests",
    "view_outward",
    "create_outward",
    "view_reports",
    "export_reports",
    "view_user_management",
    "manage_users"
]

# Admin gets all permissions by default
ADMIN_PERMISSIONS = AVAILABLE_PERMISSIONS.copy()

def get_user_permissions(user_role: str, user_permissions: List[str]) -> List[str]:
    """Get effective permissions for a user"""
    if user_role == "admin":
        return ADMIN_PERMISSIONS
    return user_permissions

def check_permission(user, required_permission: str) -> bool:
    """Check if user has a specific permission"""
    effective_permissions = get_user_permissions(user.role, user.permissions)
    return required_permission in effective_permissions
security = HTTPBearer()

# Create the main app
app = FastAPI(title="Electronics Board Inventory Management")
api_router = APIRouter(prefix="/api")

# Health check routes
@api_router.get("/")
async def root():
    return {"message": "Electronics Board Inventory API", "status": "running"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "service": "electronics_inventory_api"}

# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    designation: Optional[str] = None
    role: str = "user"  # admin, manager, user
    permissions: List[str] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True

class UserCreate(BaseModel):
    email: EmailStr
    first_name: str = Field(min_length=1, description="First name is required")
    last_name: str = Field(min_length=1, description="Last name is required")
    designation: str = Field(min_length=1, description="Designation is required")
    password: str

class UserPasswordReset(BaseModel):
    user_id: str
    new_password: str = Field(min_length=6, description="Password must be at least 6 characters")

class UserPermissionUpdate(BaseModel):
    user_id: str
    permissions: List[str]

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: str
    manufacturer: str
    version: str
    lead_time_days: int
    minimum_stock_quantity: int
    picture_url: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class CategoryCreate(BaseModel):
    name: str
    description: str
    manufacturer: str
    version: str
    lead_time_days: int
    minimum_stock_quantity: int
    picture_url: Optional[str] = None

class Board(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category_id: str
    serial_number: str
    location: str = "In stock"  # In stock, Issued for machine, Repairing, Issued for spares, At customer site
    condition: str = "OK"  # OK, Needs repair, Scrap
    issued_by: Optional[str] = None
    issued_to: Optional[str] = None
    qc_by: Optional[str] = None
    inward_date_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    issued_date_time: Optional[datetime] = None
    project_number: Optional[str] = None
    comments: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class BoardCreate(BaseModel):
    category_id: str
    serial_number: str
    location: str = "In stock"
    condition: str = "OK"
    qc_by: Optional[str] = None
    comments: Optional[str] = None

class BoardUpdate(BaseModel):
    location: Optional[str] = None
    condition: Optional[str] = None
    issued_by: Optional[str] = None
    issued_to: Optional[str] = None
    project_number: Optional[str] = None
    comments: Optional[str] = None

class BoardRequest(BaseModel):
    category_id: str
    serial_number: Optional[str] = None  # Optional - assigned during issuance
    condition: Optional[str] = None  # Optional - assigned during issuance
    quantity: Optional[int] = None  # Used when serial_number not yet assigned

class BulkIssueRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    boards: List[BoardRequest]  # All boards in this bulk request (with or without serial numbers)
    requested_by: str
    issued_to: str
    project_number: str
    comments: Optional[str] = None
    status: str = "pending"  # pending, approved, issued, rejected
    approved_by: Optional[str] = None
    created_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    approved_date: Optional[str] = None
    # User display names (populated by API)
    requested_by_name: Optional[str] = None
    issued_to_name: Optional[str] = None

class IssueRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category_id: str
    serial_number: Optional[str] = None  # Can be specific or "any available"
    requested_by: str
    issued_to: str
    project_number: str
    comments: Optional[str] = None
    status: str = "pending"  # pending, approved, issued, rejected
    request_date_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    approved_by: Optional[str] = None
    approved_date_time: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # User display names (populated by API)
    requested_by_name: Optional[str] = None
    issued_to_name: Optional[str] = None

class IssueRequestCreate(BaseModel):
    category_id: str
    serial_number: Optional[str] = None
    issued_to: str
    project_number: str
    comments: Optional[str] = None

class CategoryBoardRequest(BaseModel):
    category_id: str
    quantity: Optional[int] = Field(None, gt=0, le=50)  # Optional - either quantity or serial_numbers
    serial_numbers: Optional[List[str]] = Field(None, max_items=50)  # Specific serial numbers
    
    @validator('serial_numbers', 'quantity')
    def validate_quantity_or_serials(cls, v, values):
        # Must have either quantity OR serial_numbers, not both or neither
        if 'quantity' in values and 'serial_numbers' in values:
            if values.get('quantity') and v:
                raise ValueError('Cannot specify both quantity and serial_numbers')
            if not values.get('quantity') and not v:
                raise ValueError('Must specify either quantity or serial_numbers')
        return v

class BulkIssueRequestCreate(BaseModel):
    categories: List[CategoryBoardRequest] = Field(min_items=1, max_items=5)  # Up to 5 categories
    issued_to: str
    project_number: str
    comments: Optional[str] = None

class IssueRequestUpdate(BaseModel):
    status: str
    approved_by: Optional[str] = None
    serial_number: Optional[str] = None  # Assigned when approved

class UserUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    designation: Optional[str] = None
    role: Optional[str] = None
    permissions: Optional[List[str]] = None
    is_active: Optional[bool] = None

# Auth functions
def verify_password(plain_password, hashed_password):
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def get_password_hash(password):
    # Truncate password to 72 bytes if necessary (bcrypt limitation)
    password_bytes = password.encode('utf-8')[:72]
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password_bytes, salt).decode('utf-8')

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"email": email})
    if user is None:
        raise credentials_exception
    return User(**user)

from fastapi import Query

from fastapi.security import HTTPBearer
from typing import Optional

security_optional = HTTPBearer(auto_error=False)

async def get_current_user_flexible(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security_optional),
    token: Optional[str] = Query(None)
):
    """Auth that accepts token from header or query parameter"""
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    
    # Get token from header or query parameter
    auth_token = None
    if credentials and credentials.credentials:
        auth_token = credentials.credentials
    elif token:
        auth_token = token
    
    if not auth_token:
        raise credentials_exception
    
    try:
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"email": email})
    if user is None:
        raise credentials_exception
    return User(**user)

# Auth routes
@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate password
    if len(user_data.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    # Create user
    hashed_password = get_password_hash(user_data.password)
    user = User(
        email=user_data.email,
        first_name=user_data.first_name,
        last_name=user_data.last_name,
        designation=user_data.designation
    )
    user_dict = user.dict()
    user_dict["password"] = hashed_password
    
    await db.users.insert_one(user_dict)
    
    # Create token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    # Find user
    user_doc = await db.users.find_one({"email": user_data.email})
    if not user_doc or not verify_password(user_data.password, user_doc["password"]):
        raise HTTPException(status_code=400, detail="Incorrect email or password")
    
    user = User(**user_doc)
    
    # Create token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# Category routes
@api_router.post("/categories", response_model=Category)
async def create_category(category_data: CategoryCreate, current_user: User = Depends(get_current_user)):
    # Check if category name already exists
    existing = await db.categories.find_one({"name": category_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Category name already exists")
    
    category = Category(**category_data.dict(), created_by=current_user.email)
    await db.categories.insert_one(category.dict())
    return category

@api_router.get("/categories", response_model=List[Category])
async def get_categories(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, "view_categories"):
        raise HTTPException(status_code=403, detail="Permission denied: view_categories required")
    categories = await db.categories.find().to_list(1000)
    return [Category(**cat) for cat in categories]

@api_router.get("/categories/{category_id}", response_model=Category)
async def get_category(category_id: str, current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, "view_categories"):
        raise HTTPException(status_code=403, detail="Permission denied: view_categories required")
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return Category(**category)

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, category_data: CategoryCreate, current_user: User = Depends(get_current_user)):
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Check if new name conflicts with existing (excluding current)
    if category_data.name != category["name"]:
        existing = await db.categories.find_one({"name": category_data.name})
        if existing:
            raise HTTPException(status_code=400, detail="Category name already exists")
    
    await db.categories.update_one(
        {"id": category_id},
        {"$set": category_data.dict()}
    )
    
    updated_category = await db.categories.find_one({"id": category_id})
    return Category(**updated_category)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: User = Depends(get_current_user)):
    # Check if category has any boards
    boards_count = await db.boards.count_documents({"category_id": category_id})
    if boards_count > 0:
        raise HTTPException(status_code=400, detail="Cannot delete category with existing boards")
    
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {"message": "Category deleted successfully"}

# Board routes
@api_router.post("/boards", response_model=Board)
async def create_board(board_data: BoardCreate, current_user: User = Depends(get_current_user)):
    # Verify category exists
    category = await db.categories.find_one({"id": board_data.category_id})
    if not category:
        raise HTTPException(status_code=400, detail="Category not found")
    
    # Check if serial number exists in this category
    existing = await db.boards.find_one({
        "category_id": board_data.category_id,
        "serial_number": board_data.serial_number
    })
    if existing:
        raise HTTPException(status_code=400, detail="Serial number already exists in this category")
    
    board = Board(**board_data.dict(), created_by=current_user.email)
    await db.boards.insert_one(board.dict())
    return board

@api_router.get("/boards", response_model=List[Board])
async def get_boards(current_user: User = Depends(get_current_user)):
    boards = await db.boards.find().to_list(1000)
    return [Board(**board) for board in boards]

@api_router.get("/boards/{board_id}", response_model=Board)
async def get_board(board_id: str, current_user: User = Depends(get_current_user)):
    board = await db.boards.find_one({"id": board_id})
    if not board:
        raise HTTPException(status_code=404, detail="Board not found")
    return Board(**board)

@api_router.put("/boards/{board_id}", response_model=Board)
async def update_board(board_id: str, board_data: BoardUpdate, current_user: User = Depends(get_current_user)):
    board = await db.boards.find_one({"id": board_id})
    if not board:
        raise HTTPException(status_code=404, detail="Board not found")
    
    update_data = {k: v for k, v in board_data.dict().items() if v is not None}
    
    # If issuing the board, set issued_date_time
    if update_data.get("location") and update_data["location"] != "In stock":
        update_data["issued_date_time"] = datetime.now(timezone.utc)
        update_data["issued_by"] = current_user.email
    
    await db.boards.update_one(
        {"id": board_id},
        {"$set": update_data}
    )
    
    updated_board = await db.boards.find_one({"id": board_id})
    return Board(**updated_board)

@api_router.delete("/boards/{board_id}")
async def delete_board(board_id: str, current_user: User = Depends(get_current_user)):
    result = await db.boards.delete_one({"id": board_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Board not found")
    
    return {"message": "Board deleted successfully"}

# Search route
@api_router.get("/search")
async def search_boards(
    query: Optional[str] = None,
    category_id: Optional[str] = None,
    location: Optional[str] = None,
    condition: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    filter_conditions = {}
    
    if category_id:
        filter_conditions["category_id"] = category_id
    if location:
        filter_conditions["location"] = location
    if condition:
        filter_conditions["condition"] = condition
    
    # Text search in serial_number, issued_to, project_number, comments
    if query:
        filter_conditions["$or"] = [
            {"serial_number": {"$regex": query, "$options": "i"}},
            {"issued_to": {"$regex": query, "$options": "i"}},
            {"project_number": {"$regex": query, "$options": "i"}},
            {"comments": {"$regex": query, "$options": "i"}}
        ]
    
    boards = await db.boards.find(filter_conditions).to_list(1000)
    return [Board(**board) for board in boards]

# Issue Request routes
@api_router.post("/issue-requests", response_model=IssueRequest)
async def create_issue_request(request_data: IssueRequestCreate, current_user: User = Depends(get_current_user)):
    # Verify category exists
    category = await db.categories.find_one({"id": request_data.category_id})
    if not category:
        raise HTTPException(status_code=400, detail="Category not found")
    
    # If specific serial number requested, verify it exists and is available
    if request_data.serial_number:
        board = await db.boards.find_one({
            "category_id": request_data.category_id,
            "serial_number": request_data.serial_number,
            "$or": [
                {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                {"location": "Repairing", "condition": "Repaired"}
            ]
        })
        if not board:
            raise HTTPException(status_code=400, detail="Board not available or not found")
    
    issue_request = IssueRequest(**request_data.dict(), requested_by=current_user.email)
    await db.issue_requests.insert_one(issue_request.dict())
    return issue_request

@api_router.post("/issue-requests/bulk")
async def create_bulk_issue_request(bulk_request: BulkIssueRequestCreate, current_user: User = Depends(get_current_user)):
    """Create a single bulk issue request for multiple categories - boards assigned during approval"""
    
    board_requests = []
    failed_categories = []
    total_boards = 0
    
    # Validate all categories and check availability
    for category_request in bulk_request.categories:
        # Verify category exists
        category = await db.categories.find_one({"id": category_request.category_id})
        if not category:
            failed_categories.append({
                "category_id": category_request.category_id,
                "category_name": "Unknown", 
                "error": "Category not found"
            })
            continue
        
        # Check available boards count
        available_count = await db.boards.count_documents({
            "category_id": category_request.category_id,
            "location": "In stock",
            "condition": {"$in": ["New", "Repaired"]}
        })
        
        # Validate quantity
        if category_request.quantity:
            if available_count < category_request.quantity:
                failed_categories.append({
                    "category_id": category_request.category_id,
                    "category_name": category.get("name", "Unknown"),
                    "error": f"Not enough boards available. Requested: {category_request.quantity}, Available: {available_count}"
                })
                continue
            
            # Store category-quantity pair (boards will be assigned during issuance)
            board_requests.append(
                BoardRequest(
                    category_id=category_request.category_id,
                    quantity=category_request.quantity
                )
            )
            total_boards += category_request.quantity
    
    if failed_categories:
        # Some categories failed
        error_details = "; ".join([f"{cat['category_name']}: {cat['error']}" for cat in failed_categories])
        raise HTTPException(status_code=400, detail=f"Bulk request failed: {error_details}")
    
    if not board_requests:
        raise HTTPException(status_code=400, detail="No valid board requests")
    
    # Create bulk issue request without specific boards assigned
    bulk_issue_request = BulkIssueRequest(
        boards=board_requests,
        requested_by=current_user.email,
        issued_to=bulk_request.issued_to,
        project_number=bulk_request.project_number,
        comments=bulk_request.comments
    )
    
    await db.bulk_issue_requests.insert_one(bulk_issue_request.dict())
    
    return {
        "message": f"Successfully created bulk issue request for {total_boards} boards",
        "request_id": bulk_issue_request.id,
        "total_boards": total_boards,
        "successful": total_boards,  # For frontend compatibility
        "boards": [{"category_id": br.category_id, "quantity": br.quantity} for br in board_requests]
    }

@api_router.get("/issue-requests", response_model=List[IssueRequest])
async def get_issue_requests(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, "view_issue_requests"):
        raise HTTPException(status_code=403, detail="Permission denied: view_issue_requests required")
    
    # Users can see their own requests, admins can see all
    filter_query = {}
    if current_user.role != "admin":
        filter_query["requested_by"] = current_user.email
    
    requests = await db.issue_requests.find(filter_query).to_list(1000)
    
    # Get all users to populate names
    all_users = await db.users.find({}).to_list(1000)
    user_map = {user["email"]: f"{user.get('first_name', '')} {user.get('last_name', '')}" for user in all_users}
    
    # Enhance requests with user names
    enhanced_requests = []
    for req in requests:
        req_dict = dict(req)
        req_dict["requested_by_name"] = user_map.get(req["requested_by"], req["requested_by"])
        req_dict["issued_to_name"] = user_map.get(req["issued_to"], req["issued_to"]) if req.get("issued_to") else ""
        enhanced_requests.append(IssueRequest(**req_dict))
    
    return enhanced_requests

@api_router.get("/bulk-issue-requests", response_model=List[BulkIssueRequest])
async def get_bulk_issue_requests(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, "view_issue_requests"):
        raise HTTPException(status_code=403, detail="Permission denied: view_issue_requests required")
    
    requests = await db.bulk_issue_requests.find().to_list(length=None)
    
    # Get all users to populate names
    all_users = await db.users.find({}).to_list(1000)
    user_map = {user["email"]: f"{user.get('first_name', '')} {user.get('last_name', '')}" for user in all_users}
    
    # Enhance requests with user names
    enhanced_requests = []
    for req in requests:
        req_dict = dict(req)
        req_dict["requested_by_name"] = user_map.get(req["requested_by"], req["requested_by"])
        req_dict["issued_to_name"] = user_map.get(req["issued_to"], req["issued_to"]) if req.get("issued_to") else ""
        enhanced_requests.append(BulkIssueRequest(**req_dict))
    
    return enhanced_requests

@api_router.post("/boards/preview-auto-select")
async def preview_auto_select_boards(request: dict, current_user: User = Depends(get_current_user)):
    """Preview which boards will be auto-selected for a given category and quantity"""
    category_id = request.get("category_id")
    quantity = request.get("quantity", 0)
    
    if not category_id or quantity <= 0:
        raise HTTPException(status_code=400, detail="Category ID and quantity required")
    
    # Get available boards
    available_boards = await db.boards.find({
        "category_id": category_id,
        "location": "In stock",
        "condition": "New"
    }).to_list(None)
    
    if len(available_boards) < quantity:
        raise HTTPException(status_code=400, detail=f"Not enough boards available. Requested: {quantity}, Available: {len(available_boards)}")
    
    selected_boards = available_boards[:quantity]
    return {
        "selected_boards": [
            {
                "id": board["id"],
                "serial_number": board["serial_number"],
                "condition": board["condition"]
            } for board in selected_boards
        ],
        "total_available": len(available_boards)
    }

@api_router.put("/issue-requests/{request_id}", response_model=IssueRequest)
async def update_issue_request(request_id: str, update_data: IssueRequestUpdate, current_user: User = Depends(get_current_user)):
    # Only admins can approve/reject requests
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update issue requests")
    
    request_doc = await db.issue_requests.find_one({"id": request_id})
    if not request_doc:
        raise HTTPException(status_code=404, detail="Issue request not found")
    
    update_dict = update_data.dict(exclude_unset=True)
    if update_data.status == "approved":
        update_dict["approved_by"] = current_user.email
        update_dict["approved_date_time"] = datetime.now(timezone.utc)
    
    await db.issue_requests.update_one(
        {"id": request_id},
        {"$set": update_dict}
    )
    
    updated_request = await db.issue_requests.find_one({"id": request_id})
    return IssueRequest(**updated_request)

class BulkIssueRequestUpdate(BaseModel):
    status: str
    approved_by: Optional[str] = None

@api_router.put("/bulk-issue-requests/{request_id}", response_model=BulkIssueRequest)
async def update_bulk_issue_request(request_id: str, update_data: BulkIssueRequestUpdate, current_user: User = Depends(get_current_user)):
    # Only admins can approve/reject bulk requests
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update bulk issue requests")
    
    request_doc = await db.bulk_issue_requests.find_one({"id": request_id})
    if not request_doc:
        raise HTTPException(status_code=404, detail="Bulk issue request not found")
    
    update_dict = update_data.dict(exclude_unset=True)
    if update_data.status == "approved":
        update_dict["approved_by"] = current_user.email
        update_dict["approved_date"] = datetime.now(timezone.utc).isoformat()
    
    await db.bulk_issue_requests.update_one(
        {"id": request_id},
        {"$set": update_dict}
    )
    
    updated_request = await db.bulk_issue_requests.find_one({"id": request_id})
    return BulkIssueRequest(**updated_request)

@api_router.delete("/issue-requests/{request_id}")
async def delete_issue_request(request_id: str, current_user: User = Depends(get_current_user)):
    # Only admins can delete requests
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete issue requests")
    
    result = await db.issue_requests.delete_one({"id": request_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Issue request not found")
    
    return {"message": "Issue request deleted successfully"}

@api_router.delete("/bulk-issue-requests/{request_id}")
async def delete_bulk_issue_request(request_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.bulk_issue_requests.delete_one({"id": request_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    
    return {"message": "Bulk request deleted successfully"}

class BoardAssignmentUpdate(BaseModel):
    boards: List[BoardRequest]
    status: str = "approved"

@api_router.put("/bulk-issue-requests/{request_id}/assign-boards")
async def assign_boards_to_bulk_request(
    request_id: str, 
    assignment: BoardAssignmentUpdate,
    current_user: User = Depends(get_current_user)
):
    """Assign specific boards to a bulk request and approve it"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can assign boards and approve requests")
    
    # Find the bulk request
    request_doc = await db.bulk_issue_requests.find_one({"id": request_id})
    if not request_doc:
        raise HTTPException(status_code=404, detail="Bulk issue request not found")
    
    # Validate all boards are available
    for board_request in assignment.boards:
        if board_request.serial_number:
            board = await db.boards.find_one({
                "category_id": board_request.category_id,
                "serial_number": board_request.serial_number,
                "$or": [
                    {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                    {"location": "Repairing", "condition": "Repaired"}
                ]
            })
            if not board:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Board {board_request.serial_number} is not available"
                )
    
    # Update the bulk request with assigned boards and approved status
    update_dict = {
        "boards": [br.dict() for br in assignment.boards],
        "status": assignment.status,
        "approved_by": current_user.email,
        "approved_date": datetime.now(timezone.utc).isoformat()
    }
    
    await db.bulk_issue_requests.update_one(
        {"id": request_id},
        {"$set": update_dict}
    )
    
    updated_request = await db.bulk_issue_requests.find_one({"id": request_id})
    return BulkIssueRequest(**updated_request)

class OutwardRequest(BaseModel):
    request_id: Optional[str] = None
    board_id: Optional[str] = None
    issued_to: Optional[str] = None
    project_number: Optional[str] = None
    issued_by_override: Optional[str] = None
    issued_to_override: Optional[str] = None
    comments: Optional[str] = None

# Outward (Issue Board) route
@api_router.post("/outward")
async def issue_board(
    outward_data: OutwardRequest,
    current_user: User = Depends(get_current_user)
):
    if outward_data.request_id:
        # Check if it's an individual request first
        request_doc = await db.issue_requests.find_one({"id": outward_data.request_id})
        
        # If not found, check if it's a bulk request
        if not request_doc:
            bulk_request_doc = await db.bulk_issue_requests.find_one({"id": outward_data.request_id})
            if bulk_request_doc and bulk_request_doc["status"] == "approved":
                # Handle bulk request issuance
                issued_boards = []
                failed_boards = []
                
                for board_request in bulk_request_doc["boards"]:
                    try:
                        # Check if boards are already assigned (old format) or need assignment (new format)
                        if board_request.get("serial_number"):
                            # Old format: specific serial number already assigned
                            query = {
                                "category_id": board_request["category_id"],
                                "serial_number": board_request["serial_number"],
                                "$or": [
                                    {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                                    {"location": "Repairing", "condition": "Repaired"}
                                ]
                            }
                            
                            board = await db.boards.find_one(query)
                            if board:
                                # Update board
                                await db.boards.update_one(
                                    {"id": board["id"]},
                                    {"$set": {
                                        "location": "Issued for machine",
                                        "issued_by": outward_data.issued_by_override or current_user.email,
                                        "issued_to": outward_data.issued_to_override or bulk_request_doc["issued_to"],
                                        "project_number": bulk_request_doc["project_number"],
                                        "issued_date_time": datetime.now(timezone.utc),
                                        "comments": outward_data.comments or ""
                                    }}
                                )
                                issued_boards.append(board["serial_number"])
                            else:
                                failed_boards.append(f"Category: {board_request['category_id']}, Serial: {board_request['serial_number']}")
                        elif board_request.get("quantity"):
                            # New format: assign boards now based on quantity
                            quantity = board_request["quantity"]
                            
                            # Find available boards for this category
                            available_boards = await db.boards.find({
                                "category_id": board_request["category_id"],
                                "$or": [
                                    {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                                    {"location": "Repairing", "condition": "Repaired"}
                                ]
                            }).to_list(quantity)
                            
                            if len(available_boards) < quantity:
                                failed_boards.append(f"Category: {board_request['category_id']} - Requested {quantity}, only {len(available_boards)} available")
                                continue
                            
                            # Issue the boards
                            for board in available_boards[:quantity]:
                                await db.boards.update_one(
                                    {"id": board["id"]},
                                    {"$set": {
                                        "location": "Issued for machine",
                                        "issued_by": outward_data.issued_by_override or current_user.email,
                                        "issued_to": outward_data.issued_to_override or bulk_request_doc["issued_to"],
                                        "project_number": bulk_request_doc["project_number"],
                                        "issued_date_time": datetime.now(timezone.utc),
                                        "comments": outward_data.comments or ""
                                    }}
                                )
                                issued_boards.append(board["serial_number"])
                        
                    except Exception as e:
                        failed_boards.append(f"Category: {board_request['category_id']} - Error: {str(e)}")
                
                # Calculate expected total boards
                expected_total = sum([
                    br.get("quantity", 1) if br.get("quantity") else 1 
                    for br in bulk_request_doc["boards"]
                ])
                
                # Update bulk request status
                if len(issued_boards) == expected_total:
                    status = "issued"
                elif len(issued_boards) > 0:
                    status = "partially_issued"
                else:
                    status = "approved"  # Keep as approved if nothing was issued
                
                await db.bulk_issue_requests.update_one(
                    {"id": outward_data.request_id},
                    {"$set": {"status": status}}
                )
                
                if len(issued_boards) > 0:
                    message = f"Bulk request processed: {len(issued_boards)} boards issued"
                    if failed_boards:
                        message += f", {len(failed_boards)} failed"
                    return {"message": message, "issued_boards": issued_boards, "failed_boards": failed_boards}
                else:
                    raise HTTPException(status_code=400, detail=f"No boards could be issued. Failed: {failed_boards}")
            else:
                raise HTTPException(status_code=400, detail="Request not found or not approved")
        
        # Handle individual request (original logic)
        if request_doc["status"] != "approved":
            raise HTTPException(status_code=400, detail="Individual request not approved")
        
        # Find available board (New/Repaired boards that are In stock OR Repaired boards that are being repaired)
        query = {
            "category_id": request_doc["category_id"], 
            "$or": [
                {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                {"location": "Repairing", "condition": "Repaired"}
            ]
        }
        if request_doc.get("serial_number"):
            query["serial_number"] = request_doc["serial_number"]
        
        board = await db.boards.find_one(query)
        if not board:
            raise HTTPException(status_code=400, detail="No available board found")
        
        # Update board
        await db.boards.update_one(
            {"id": board["id"]},
            {"$set": {
                "location": "Issued for machine",
                "issued_by": outward_data.issued_by_override or current_user.email,
                "issued_to": outward_data.issued_to_override or request_doc["issued_to"],
                "project_number": request_doc["project_number"],
                "issued_date_time": datetime.now(timezone.utc),
                "comments": outward_data.comments or ""
            }}
        )
        
        # Update request status
        await db.issue_requests.update_one(
            {"id": outward_data.request_id},
            {"$set": {"status": "issued", "serial_number": board["serial_number"]}}
        )
        
        return {"message": "Board issued successfully", "serial_number": board["serial_number"]}
    
    elif outward_data.board_id and outward_data.issued_to:
        # Direct board issue (New/Repaired boards that are In stock OR Repaired boards that are being repaired)
        board = await db.boards.find_one({
            "id": outward_data.board_id,
            "$or": [
                {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                {"location": "Repairing", "condition": "Repaired"}
            ]
        })
        if not board:
            raise HTTPException(status_code=400, detail="Board not available")
        
        await db.boards.update_one(
            {"id": outward_data.board_id},
            {"$set": {
                "location": "Issued for machine",
                "issued_by": outward_data.issued_by_override or current_user.email,
                "issued_to": outward_data.issued_to_override or outward_data.issued_to,
                "project_number": outward_data.project_number or "",
                "issued_date_time": datetime.now(timezone.utc),
                "comments": outward_data.comments or ""
            }}
        )
        
        return {"message": "Board issued successfully", "serial_number": board["serial_number"]}
    
    else:
        raise HTTPException(status_code=400, detail="Either request_id or board_id with issued_to is required")

# User Management routes (Admin only)
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view users")
    
    users = await db.users.find().to_list(1000)
    return [User(**user) for user in users]

@api_router.put("/users/{user_email}", response_model=User)
async def update_user(user_email: str, update_data: UserUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update users")
    
    user = await db.users.find_one({"email": user_email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_dict = update_data.dict(exclude_unset=True)
    await db.users.update_one(
        {"email": user_email},
        {"$set": update_dict}
    )
    
    updated_user = await db.users.find_one({"email": user_email})
    return User(**updated_user)

@api_router.post("/users/reset-password")
async def reset_user_password(reset_data: UserPasswordReset, current_user: User = Depends(get_current_user)):
    """Admin endpoint to reset user password"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can reset passwords")
    
    user = await db.users.find_one({"id": reset_data.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Hash new password
    hashed_password = get_password_hash(reset_data.new_password)
    
    # Update password
    await db.users.update_one(
        {"id": reset_data.user_id},
        {"$set": {"password": hashed_password}}
    )
    
    return {"message": f"Password reset successful for user: {user['email']}"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Admin endpoint to delete user"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    # Prevent admin from deleting themselves
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Delete user
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": f"User {user['email']} deleted successfully"}

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, permission_data: UserPermissionUpdate, current_user: User = Depends(get_current_user)):
    """Admin endpoint to update user permissions"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can manage permissions")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Validate permissions
    invalid_permissions = [p for p in permission_data.permissions if p not in AVAILABLE_PERMISSIONS]
    if invalid_permissions:
        raise HTTPException(status_code=400, detail=f"Invalid permissions: {invalid_permissions}")
    
    # Update permissions
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"permissions": permission_data.permissions}}
    )
    
    updated_user = await db.users.find_one({"id": user_id})
    return {"message": f"Permissions updated for {user['email']}", "user": User(**updated_user)}

@api_router.get("/permissions/available")
async def get_available_permissions(current_user: User = Depends(get_current_user)):
    """Get list of available permissions"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view available permissions")
    
    return {"permissions": AVAILABLE_PERMISSIONS}

@api_router.get("/users/me/permissions")
async def get_my_permissions(current_user: User = Depends(get_current_user)):
    """Get current user's effective permissions"""
    effective_permissions = get_user_permissions(current_user.role, current_user.permissions)
    return {"permissions": effective_permissions}

# Reports endpoints
@api_router.get("/reports/low-stock")
async def get_low_stock_report(current_user: User = Depends(get_current_user)):
    """Get categories with stock below minimum threshold"""
    if not check_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: view_reports required")
    
    categories = await db.categories.find().to_list(1000)
    low_stock_report = []
    
    for category in categories:
        # Count current stock (In stock + Repairing with condition Repaired)
        current_stock = await db.boards.count_documents({
            "category_id": category["id"],
            "$or": [
                {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                {"location": "Repairing", "condition": "Repaired"}
            ]
        })
        
        min_stock = category.get("minimum_stock_quantity", 0)
        if current_stock < min_stock:
            shortage = min_stock - current_stock
            low_stock_report.append({
                "category_id": category["id"],
                "category_name": category["name"],
                "manufacturer": category["manufacturer"],
                "version": category["version"],
                "current_stock": current_stock,
                "minimum_stock_quantity": min_stock,
                "shortage_quantity": shortage,
                "required_to_refill": shortage
            })
    
    return low_stock_report

@api_router.get("/reports/under-repair")
async def get_under_repair_report(current_user: User = Depends(get_current_user)):
    """Get all boards that are currently under repair"""
    if not check_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: view_reports required")
    
    # Find all boards with condition "Under repair" or location "Repairing" with condition not "Repaired"
    under_repair_boards = await db.boards.find({
        "$or": [
            {"condition": "Under repair"},
            {"location": "Repairing", "condition": {"$ne": "Repaired"}}
        ]
    }).to_list(1000)
    
    # Get category information for each board
    categories = await db.categories.find().to_list(1000)
    category_map = {cat["id"]: cat for cat in categories}
    
    repair_report = []
    for board in under_repair_boards:
        category = category_map.get(board["category_id"], {})
        repair_report.append({
            "serial_number": board["serial_number"],
            "category_name": category.get("name", "Unknown"),
            "manufacturer": category.get("manufacturer", "Unknown"),
            "version": category.get("version", "Unknown"),
            "condition": board["condition"],
            "location": board["location"],
            "inward_date": board.get("inward_date_time", ""),
            "comments": board.get("comments", "")
        })
    
    return repair_report

@api_router.get("/reports/serial-history/{serial_number}")
async def get_serial_history(serial_number: str, current_user: User = Depends(get_current_user)):
    """Get complete history of a specific serial number"""
    if not check_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: view_reports required")
    
    # Find the board
    board = await db.boards.find_one({"serial_number": serial_number})
    if not board:
        raise HTTPException(status_code=404, detail="Serial number not found")
    
    # Get category info
    category = await db.categories.find_one({"id": board["category_id"]})
    
    # Get issue requests for this board
    issue_requests = await db.issue_requests.find({
        "$or": [
            {"serial_number": serial_number},
            {"category_id": board["category_id"], "serial_number": None}  # Generic requests that might have been fulfilled by this board
        ]
    }).to_list(100)
    
    # Get bulk issue requests that might include this board
    bulk_requests = await db.bulk_issue_requests.find({
        "boards.serial_number": serial_number
    }).to_list(100)
    
    # Convert MongoDB documents to proper format
    issue_history = []
    for req in issue_requests:
        # Remove MongoDB ObjectId fields and convert to proper format
        clean_req = {k: v for k, v in req.items() if k != '_id'}
        # Convert datetime objects to strings
        if 'request_date_time' in clean_req and clean_req['request_date_time']:
            clean_req['request_date_time'] = clean_req['request_date_time'].isoformat() if hasattr(clean_req['request_date_time'], 'isoformat') else str(clean_req['request_date_time'])
        if 'approved_date_time' in clean_req and clean_req['approved_date_time']:
            clean_req['approved_date_time'] = clean_req['approved_date_time'].isoformat() if hasattr(clean_req['approved_date_time'], 'isoformat') else str(clean_req['approved_date_time'])
        if 'created_at' in clean_req and clean_req['created_at']:
            clean_req['created_at'] = clean_req['created_at'].isoformat() if hasattr(clean_req['created_at'], 'isoformat') else str(clean_req['created_at'])
        issue_history.append(clean_req)
    
    bulk_history = []
    for req in bulk_requests:
        # Remove MongoDB ObjectId fields and convert to proper format
        clean_req = {k: v for k, v in req.items() if k != '_id'}
        bulk_history.append(clean_req)
    
    # Clean board details
    clean_board = {k: v for k, v in board.items() if k != '_id'}
    # Convert datetime objects to strings
    if 'inward_date_time' in clean_board and clean_board['inward_date_time']:
        clean_board['inward_date_time'] = clean_board['inward_date_time'].isoformat() if hasattr(clean_board['inward_date_time'], 'isoformat') else str(clean_board['inward_date_time'])
    if 'issued_date_time' in clean_board and clean_board['issued_date_time']:
        clean_board['issued_date_time'] = clean_board['issued_date_time'].isoformat() if hasattr(clean_board['issued_date_time'], 'isoformat') else str(clean_board['issued_date_time'])
    if 'created_at' in clean_board and clean_board['created_at']:
        clean_board['created_at'] = clean_board['created_at'].isoformat() if hasattr(clean_board['created_at'], 'isoformat') else str(clean_board['created_at'])

    history = {
        "serial_number": serial_number,
        "category_name": category["name"] if category else "Unknown",
        "manufacturer": category["manufacturer"] if category else "Unknown",
        "version": category["version"] if category else "Unknown",
        "current_status": {
            "condition": board["condition"],
            "location": board["location"],
            "issued_to": board.get("issued_to", ""),
            "issued_by": board.get("issued_by", ""),
            "project_number": board.get("project_number", ""),
            "issued_date": str(board.get("issued_date_time", "")) if board.get("issued_date_time") else "",
            "inward_date": str(board.get("inward_date_time", "")) if board.get("inward_date_time") else "",
            "comments": board.get("comments", "")
        },
        "issue_history": issue_history,
        "bulk_request_history": bulk_history,
        "board_details": clean_board
    }
    
    return history

@api_router.get("/reports/serial-numbers/{category_id}")
async def get_serial_numbers_by_category(category_id: str, current_user: User = Depends(get_current_user)):
    """Get all serial numbers for a specific category"""
    if not check_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: view_reports required")
    
    # Verify category exists
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Get all boards for this category
    boards = await db.boards.find({"category_id": category_id}).to_list(1000)
    
    serial_numbers = []
    for board in boards:
        serial_numbers.append({
            "serial_number": board["serial_number"],
            "condition": board["condition"],
            "location": board["location"],
            "status": f"{board['condition']} - {board['location']}"
        })
    
    # Sort by serial number
    serial_numbers.sort(key=lambda x: x["serial_number"])
    
    return {
        "category_id": category_id,
        "category_name": category["name"],
        "serial_numbers": serial_numbers
    }

@api_router.get("/reports/category-export/{category_id}")
async def get_category_export_data(category_id: str, current_user: User = Depends(get_current_user)):
    """Get all data for a specific category including boards and related requests"""
    if not check_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: view_reports required")
    
    # Get category info
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Get all boards for this category
    boards = await db.boards.find({"category_id": category_id}).to_list(1000)
    
    # Get all issue requests for this category
    issue_requests = await db.issue_requests.find({"category_id": category_id}).to_list(1000)
    
    # Get bulk issue requests that include boards from this category
    bulk_requests = await db.bulk_issue_requests.find({
        "boards.category_id": category_id
    }).to_list(1000)
    
    # Clean category data
    clean_category = {k: v for k, v in category.items() if k != '_id'}
    if 'created_at' in clean_category and clean_category['created_at']:
        clean_category['created_at'] = clean_category['created_at'].isoformat() if hasattr(clean_category['created_at'], 'isoformat') else str(clean_category['created_at'])
    
    # Clean boards data
    clean_boards = []
    for board in boards:
        clean_board = {k: v for k, v in board.items() if k != '_id'}
        # Convert datetime objects to strings
        if 'inward_date_time' in clean_board and clean_board['inward_date_time']:
            clean_board['inward_date_time'] = clean_board['inward_date_time'].isoformat() if hasattr(clean_board['inward_date_time'], 'isoformat') else str(clean_board['inward_date_time'])
        if 'issued_date_time' in clean_board and clean_board['issued_date_time']:
            clean_board['issued_date_time'] = clean_board['issued_date_time'].isoformat() if hasattr(clean_board['issued_date_time'], 'isoformat') else str(clean_board['issued_date_time'])
        if 'created_at' in clean_board and clean_board['created_at']:
            clean_board['created_at'] = clean_board['created_at'].isoformat() if hasattr(clean_board['created_at'], 'isoformat') else str(clean_board['created_at'])
        clean_boards.append(clean_board)
    
    # Clean issue requests data
    clean_issue_requests = []
    for req in issue_requests:
        clean_req = {k: v for k, v in req.items() if k != '_id'}
        # Convert datetime objects to strings
        if 'request_date_time' in clean_req and clean_req['request_date_time']:
            clean_req['request_date_time'] = clean_req['request_date_time'].isoformat() if hasattr(clean_req['request_date_time'], 'isoformat') else str(clean_req['request_date_time'])
        if 'approved_date_time' in clean_req and clean_req['approved_date_time']:
            clean_req['approved_date_time'] = clean_req['approved_date_time'].isoformat() if hasattr(clean_req['approved_date_time'], 'isoformat') else str(clean_req['approved_date_time'])
        if 'created_at' in clean_req and clean_req['created_at']:
            clean_req['created_at'] = clean_req['created_at'].isoformat() if hasattr(clean_req['created_at'], 'isoformat') else str(clean_req['created_at'])
        clean_issue_requests.append(clean_req)
    
    # Clean bulk requests data
    clean_bulk_requests = []
    for req in bulk_requests:
        clean_req = {k: v for k, v in req.items() if k != '_id'}
        clean_bulk_requests.append(clean_req)

    export_data = {
        "category": clean_category,
        "boards": clean_boards,
        "issue_requests": clean_issue_requests,
        "bulk_issue_requests": clean_bulk_requests,
        "statistics": {
            "total_boards": len(boards),
            "in_stock": len([b for b in boards if b["location"] == "In stock"]),
            "issued": len([b for b in boards if b["location"] == "Issued for machine"]),
            "repairing": len([b for b in boards if b["location"] == "Repairing"]),
            "total_requests": len(issue_requests),
            "total_bulk_requests": len(bulk_requests)
        }
    }
    
    return export_data

# Excel Export endpoints
@api_router.get("/reports/export/low-stock")
async def export_low_stock_excel(current_user: User = Depends(get_current_user_flexible)):
    """Export low stock report as Excel file"""
    if not check_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: export_reports required")
    
    # Get low stock data
    categories = await db.categories.find().to_list(1000)
    low_stock_data = []
    
    for category in categories:
        current_stock = await db.boards.count_documents({
            "category_id": category["id"],
            "$or": [
                {"location": "In stock", "condition": {"$in": ["New", "Repaired"]}},
                {"location": "Repairing", "condition": "Repaired"}
            ]
        })
        
        min_stock = category.get("minimum_stock_quantity", 0)
        if current_stock < min_stock:
            shortage = min_stock - current_stock
            low_stock_data.append({
                "Category Name": category["name"],
                "Manufacturer": category["manufacturer"],
                "Version": category["version"],
                "Current Stock": current_stock,
                "Minimum Required": min_stock,
                "Shortage": shortage,
                "Required to Refill": shortage
            })
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Low Stock Report"
    
    # Add headers
    headers = list(low_stock_data[0].keys()) if low_stock_data else ["No Data"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    for row, data in enumerate(low_stock_data, 2):
        for col, value in enumerate(data.values(), 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"low_stock_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\"",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Description": "File Transfer",
        "Content-Transfer-Encoding": "binary",
        "Cache-Control": "must-revalidate, post-check=0, pre-check=0",
        "Pragma": "public",
        "Expires": "0",
        "Access-Control-Expose-Headers": "Content-Disposition, Content-Type"
    }
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@api_router.get("/reports/export/under-repair")
async def export_under_repair_excel(current_user: User = Depends(get_current_user_flexible)):
    """Export under repair report as Excel file"""
    if not check_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: export_reports required")
    
    # Get under repair data
    under_repair_boards = await db.boards.find({
        "$or": [
            {"condition": "Under repair"},
            {"location": "Repairing", "condition": {"$ne": "Repaired"}}
        ]
    }).to_list(1000)
    
    categories = await db.categories.find().to_list(1000)
    category_map = {cat["id"]: cat for cat in categories}
    
    repair_data = []
    for board in under_repair_boards:
        category = category_map.get(board["category_id"], {})
        repair_data.append({
            "Serial Number": board["serial_number"],
            "Category": category.get("name", "Unknown"),
            "Manufacturer": category.get("manufacturer", "Unknown"),
            "Version": category.get("version", "Unknown"),
            "Condition": board["condition"],
            "Location": board["location"],
            "Inward Date": board.get("inward_date_time", ""),
            "Comments": board.get("comments", "")
        })
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Under Repair Report"
    
    # Add headers
    headers = list(repair_data[0].keys()) if repair_data else ["No Data"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    for row, data in enumerate(repair_data, 2):
        for col, value in enumerate(data.values(), 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"under_repair_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\"",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Description": "File Transfer",
        "Content-Transfer-Encoding": "binary",
        "Cache-Control": "must-revalidate, post-check=0, pre-check=0",
        "Pragma": "public",
        "Expires": "0",
        "Access-Control-Expose-Headers": "Content-Disposition, Content-Type"
    }
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@api_router.get("/reports/export/serial-history/{serial_number}")
async def export_serial_history_excel(serial_number: str, current_user: User = Depends(get_current_user_flexible)):
    """Export serial number history as Excel file"""
    if not check_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: export_reports required")
    
    # Get serial history data
    board = await db.boards.find_one({"serial_number": serial_number})
    if not board:
        raise HTTPException(status_code=404, detail="Serial number not found")
    
    category = await db.categories.find_one({"id": board["category_id"]})
    
    # Create Excel file with multiple sheets
    wb = openpyxl.Workbook()
    
    # Sheet 1: Board Details
    ws1 = wb.active
    ws1.title = "Board Details"
    
    board_details = [
        ["Serial Number", board["serial_number"]],
        ["Category", category["name"] if category else "Unknown"],
        ["Manufacturer", category["manufacturer"] if category else "Unknown"],
        ["Version", category["version"] if category else "Unknown"],
        ["Current Condition", board["condition"]],
        ["Current Location", board["location"]],
        ["Issued To", board.get("issued_to", "")],
        ["Issued By", board.get("issued_by", "")],
        ["Project Number", board.get("project_number", "")],
        ["Issue Date", str(board.get("issued_date_time", ""))],
        ["Inward Date", str(board.get("inward_date_time", ""))],
        ["Comments", board.get("comments", "")]
    ]
    
    for row, (key, value) in enumerate(board_details, 1):
        ws1.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=value)
    
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 30
    
    # Sheet 2: Issue Requests History
    ws2 = wb.create_sheet("Issue History")
    issue_requests = await db.issue_requests.find({
        "$or": [
            {"serial_number": serial_number},
            {"category_id": board["category_id"], "serial_number": None}
        ]
    }).to_list(100)
    
    if issue_requests:
        headers = ["Request ID", "Status", "Requested By", "Issued To", "Project Number", "Created Date", "Comments"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, request in enumerate(issue_requests, 2):
            ws2.cell(row=row, column=1, value=request.get("id", ""))
            ws2.cell(row=row, column=2, value=request.get("status", ""))
            ws2.cell(row=row, column=3, value=request.get("requested_by", ""))
            ws2.cell(row=row, column=4, value=request.get("issued_to", ""))
            ws2.cell(row=row, column=5, value=request.get("project_number", ""))
            ws2.cell(row=row, column=6, value=str(request.get("created_date_time", "")))
            ws2.cell(row=row, column=7, value=request.get("comments", ""))
    else:
        ws2.cell(row=1, column=1, value="No issue history found")
    
    # Auto-adjust columns for sheet 2
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"serial_history_{serial_number}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\"",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Description": "File Transfer",
        "Content-Transfer-Encoding": "binary",
        "Cache-Control": "must-revalidate, post-check=0, pre-check=0",
        "Pragma": "public",
        "Expires": "0",
        "Access-Control-Expose-Headers": "Content-Disposition, Content-Type"
    }
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@api_router.get("/reports/export/category/{category_id}")
async def export_category_excel(category_id: str, current_user: User = Depends(get_current_user_flexible)):
    """Export complete category data as Excel file"""
    if not check_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="Permission denied: export_reports required")
    
    # Get category data
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    boards = await db.boards.find({"category_id": category_id}).to_list(1000)
    issue_requests = await db.issue_requests.find({"category_id": category_id}).to_list(1000)
    bulk_requests = await db.bulk_issue_requests.find({
        "boards.category_id": category_id
    }).to_list(1000)
    
    # Create Excel file with multiple sheets
    wb = openpyxl.Workbook()
    
    # Sheet 1: Category Info
    ws1 = wb.active
    ws1.title = "Category Info"
    
    category_info = [
        ["Category ID", category["id"]],
        ["Name", category["name"]],
        ["Description", category["description"]],
        ["Manufacturer", category["manufacturer"]],
        ["Version", category["version"]],
        ["Lead Time Days", category["lead_time_days"]],
        ["Minimum Stock Quantity", category["minimum_stock_quantity"]],
        ["Created Date", str(category.get("created_at", ""))],
        ["Created By", category.get("created_by", "")]
    ]
    
    for row, (key, value) in enumerate(category_info, 1):
        ws1.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=value)
    
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 30
    
    # Sheet 2: All Boards
    ws2 = wb.create_sheet("All Boards")
    if boards:
        headers = ["Serial Number", "Condition", "Location", "Issued To", "Issued By", "Project Number", "Issue Date", "Inward Date", "Comments"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, board in enumerate(boards, 2):
            ws2.cell(row=row, column=1, value=board.get("serial_number", ""))
            ws2.cell(row=row, column=2, value=board.get("condition", ""))
            ws2.cell(row=row, column=3, value=board.get("location", ""))
            ws2.cell(row=row, column=4, value=board.get("issued_to", ""))
            ws2.cell(row=row, column=5, value=board.get("issued_by", ""))
            ws2.cell(row=row, column=6, value=board.get("project_number", ""))
            ws2.cell(row=row, column=7, value=str(board.get("issued_date_time", "")))
            ws2.cell(row=row, column=8, value=str(board.get("inward_date_time", "")))
            ws2.cell(row=row, column=9, value=board.get("comments", ""))
    
    # Sheet 3: Issue Requests
    ws3 = wb.create_sheet("Issue Requests")
    if issue_requests:
        headers = ["Request ID", "Serial Number", "Status", "Requested By", "Issued To", "Project Number", "Created Date", "Comments"]
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, request in enumerate(issue_requests, 2):
            ws3.cell(row=row, column=1, value=request.get("id", ""))
            ws3.cell(row=row, column=2, value=request.get("serial_number", ""))
            ws3.cell(row=row, column=3, value=request.get("status", ""))
            ws3.cell(row=row, column=4, value=request.get("requested_by", ""))
            ws3.cell(row=row, column=5, value=request.get("issued_to", ""))
            ws3.cell(row=row, column=6, value=request.get("project_number", ""))
            ws3.cell(row=row, column=7, value=str(request.get("created_date_time", "")))
            ws3.cell(row=row, column=8, value=request.get("comments", ""))
    
    # Auto-adjust column widths for all sheets
    for ws in [ws2, ws3]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"category_{category['name']}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\"",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Description": "File Transfer",
        "Content-Transfer-Encoding": "binary",
        "Cache-Control": "must-revalidate, post-check=0, pre-check=0",
        "Pragma": "public",
        "Expires": "0",
        "Access-Control-Expose-Headers": "Content-Disposition, Content-Type"
    }
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

# Admin setup route (temporary - for initial admin creation)
@api_router.post("/setup-admin")
async def setup_admin(email: str):
    """Temporary endpoint to make the first admin user"""
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"email": email},
        {"$set": {"role": "admin", "permissions": [
            "add_category", "add_serial_number", "issue_card", 
            "raise_issue_request", "inward_new_cards", "inward_repair_cards", 
            "change_location", "change_condition"
        ]}}
    )
    
    return {"message": f"User {email} has been made an admin"}

# Dashboard stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, "view_dashboard"):
        raise HTTPException(status_code=403, detail="Permission denied: view_dashboard required")
    total_categories = await db.categories.count_documents({})
    total_boards = await db.boards.count_documents({})
    in_stock = await db.boards.count_documents({
        "location": "In stock", 
        "condition": {"$in": ["New", "Repaired"]}
    })
    issued = await db.boards.count_documents({"location": {"$ne": "In stock"}})
    repaired = await db.boards.count_documents({"condition": "Repaired"})
    scrap = await db.boards.count_documents({"condition": "Scrap"})
    pending_requests = await db.issue_requests.count_documents({"status": "pending"})
    
    return {
        "total_categories": total_categories,
        "total_boards": total_boards,
        "in_stock": in_stock,
        "issued": issued,
        "repaired": repaired,
        "scrap": scrap,
        "pending_requests": pending_requests
    }

# Include router
app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
